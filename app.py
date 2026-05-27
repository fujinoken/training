import streamlit as st
import pandas as pd
import sqlite3
from pathlib import Path
from datetime import date, datetime
import altair as alt
import shutil
import zipfile
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


APP_TITLE = "2026 年間研修管理システム Ver1.9.4 おすすめ予定回数一致版"
DB_PATH = Path("training_management.db")
UPLOAD_DIR = Path("training_uploads")
CASE_DIR = UPLOAD_DIR / "case_materials"
REPORT_DIR = UPLOAD_DIR / "staff_reports"
AUDIT_DIR = Path("audit_exports")

for p in [UPLOAD_DIR, CASE_DIR, REPORT_DIR, AUDIT_DIR]:
    p.mkdir(parents=True, exist_ok=True)

TRAINING_MASTER = [
    # 写真の年間研修表に合わせ、委員会・研修・訓練を分けて管理する
    # theme, committee, frequency, required_count, is_committee
    ("感染症の予防｜感染対策委員会", "感染対策委員会", "年2回以上 対策検討委員会開催", 2, 1),
    ("感染症の予防｜研修", "", "年1回以上研修・新規採用時研修", 1, 0),
    ("感染症の予防｜訓練", "", "年1回以上の訓練", 1, 0),
    ("身体拘束適正化｜委員会", "身体拘束の委員会", "年4回以上委員会開催", 4, 1),
    ("身体拘束適正化｜研修", "", "年2回以上研修・新規採用時研修", 2, 0),
    ("利用者の人権擁護、虐待防止｜委員会", "虐待防止委員会", "年4回以上委員会開催", 4, 1),
    ("利用者の人権擁護、虐待防止｜研修", "", "年2回以上研修・新規採用時研修", 2, 0),
    ("ハラスメント防止", "", "年2回以上研修・新規採用時研修", 2, 0),
    ("業務計画の研修｜感染症研修", "感染症", "研修 年2回以上", 2, 0),
    ("業務計画の研修｜感染症訓練", "感染症", "訓練 年2回以上", 2, 0),
    ("業務計画の研修｜自然災害研修", "自然災害", "研修 年2回以上", 2, 0),
    ("業務計画の研修｜自然災害訓練", "自然災害", "訓練 年2回以上", 2, 0),
    ("事故防止", "", "年2回", 2, 0),
    ("認知症専門ケア研修｜委員会", "", "年4回以上委員会開催", 4, 1),
    ("認知症専門ケア研修｜研修", "", "年2回以上研修・新規採用時研修", 2, 0),
    ("避難訓練", "", "年2回（防火実務研修 含む）", 2, 0),
    ("看取り", "", "年1回", 1, 0),
    ("運営推進会議", "", "年6回", 6, 0),
]
MONTHS = ["4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月", "1月", "2月", "3月"]
MONTH_NUM = {"4月":4, "5月":5, "6月":6, "7月":7, "8月":8, "9月":9, "10月":10, "11月":11, "12月":12, "1月":1, "2月":2, "3月":3}

RECOMMENDED_SCHEDULE_6_START = [
    # 年間必要回数と一致するように、6月開始で43件を分散配置
    ("6月", "2026-06-10", "感染症の予防｜感染対策委員会", "感染対策委員会。夏前に感染対策を確認。"),
    ("6月", "2026-06-17", "運営推進会議", "第1回運営推進会議。年度方針と状況共有。"),
    ("6月", "2026-06-24", "事故防止", "転倒・ヒヤリハット事例を使い、夜間対応も確認。"),

    ("7月", "2026-07-08", "感染症の予防｜研修", "感染症予防の基本確認。新規採用時研修にも使用可能。"),
    ("7月", "2026-07-22", "身体拘束適正化｜委員会", "身体拘束委員会。不適切ケアのグレー事例を確認。"),

    ("8月", "2026-08-12", "業務計画の研修｜感染症研修", "感染症BCP研修。流行期前に役割を確認。"),
    ("8月", "2026-08-19", "運営推進会議", "第2回運営推進会議。地域・家族への状況共有。"),
    ("8月", "2026-08-26", "業務計画の研修｜感染症訓練", "感染症BCP訓練。発生時の動きを確認。"),

    ("9月", "2026-09-09", "利用者の人権擁護、虐待防止｜委員会", "虐待防止委員会。声かけ・不適切ケアを確認。"),
    ("9月", "2026-09-23", "利用者の人権擁護、虐待防止｜研修", "心理的虐待・不適切ケアの事例検討。"),

    ("10月", "2026-10-07", "認知症専門ケア研修｜委員会", "認知症ケア委員会。支援方針の確認。"),
    ("10月", "2026-10-14", "身体拘束適正化｜委員会", "身体拘束委員会。前半の振り返り。"),
    ("10月", "2026-10-21", "避難訓練", "秋の避難訓練。夜間想定も確認。"),
    ("10月", "2026-10-21", "運営推進会議", "第3回運営推進会議。避難訓練等の状況共有。"),
    ("10月", "2026-10-28", "業務計画の研修｜自然災害訓練", "自然災害BCP訓練。避難・連絡体制を確認。"),

    ("11月", "2026-11-04", "利用者の人権擁護、虐待防止｜委員会", "虐待防止委員会。委員会記録の確認。"),
    ("11月", "2026-11-11", "ハラスメント防止", "職員間・利用者家族対応の基本確認。"),
    ("11月", "2026-11-18", "認知症専門ケア研修｜研修", "認知症ケア研修。声かけ・不安軽減の支援。"),
    ("11月", "2026-11-25", "身体拘束適正化｜研修", "身体拘束適正化研修。新規採用時研修にも使用可能。"),

    ("12月", "2026-12-02", "認知症専門ケア研修｜委員会", "認知症ケア委員会。困難事例の共有。"),
    ("12月", "2026-12-09", "感染症の予防｜感染対策委員会", "冬の感染症流行前の再確認。"),
    ("12月", "2026-12-16", "業務計画の研修｜感染症研修", "感染症BCP研修。冬季対応の確認。"),
    ("12月", "2026-12-16", "運営推進会議", "第4回運営推進会議。冬季対応の共有。"),
    ("12月", "2026-12-23", "感染症の予防｜訓練", "感染症発生時の訓練。"),

    ("1月", "2027-01-06", "事故防止", "年度後半の事故防止確認。ヒヤリハットの振り返り。"),
    ("1月", "2027-01-13", "身体拘束適正化｜委員会", "身体拘束委員会。年度後半の振り返り。"),
    ("1月", "2027-01-20", "業務計画の研修｜感染症訓練", "感染症BCP訓練。発生時対応の再確認。"),
    ("1月", "2027-01-27", "利用者の人権擁護、虐待防止｜委員会", "虐待防止委員会。年度後半の再確認。"),

    ("2月", "2027-02-03", "利用者の人権擁護、虐待防止｜研修", "虐待防止研修。不適切ケアの再確認。"),
    ("2月", "2027-02-10", "業務計画の研修｜自然災害研修", "自然災害BCPと備蓄・連絡体制確認。"),
    ("2月", "2027-02-17", "身体拘束適正化｜研修", "身体拘束適正化研修。年度末確認。"),
    ("2月", "2027-02-17", "認知症専門ケア研修｜委員会", "認知症ケア委員会。年度末に向けた確認。"),
    ("2月", "2027-02-18", "運営推進会議", "第5回運営推進会議。年度末前の状況共有。"),
    ("2月", "2027-02-24", "業務計画の研修｜自然災害訓練", "自然災害BCP訓練。避難・連絡体制を確認。"),
    ("2月", "2027-02-25", "避難訓練", "年度内2回目の避難訓練。防火実務研修を含む。"),

    ("3月", "2027-03-03", "身体拘束適正化｜委員会", "身体拘束委員会。年度末の記録確認。"),
    ("3月", "2027-03-10", "業務計画の研修｜自然災害研修", "自然災害BCP研修。年度末の振り返り。"),
    ("3月", "2027-03-10", "看取り", "看取り期の本人・家族支援の基本確認。"),
    ("3月", "2027-03-11", "ハラスメント防止", "ハラスメント防止研修。年度末確認。"),
    ("3月", "2027-03-17", "利用者の人権擁護、虐待防止｜委員会", "虐待防止委員会。年度末の記録確認。"),
    ("3月", "2027-03-18", "運営推進会議", "第6回運営推進会議。年度末の報告。"),
    ("3月", "2027-03-24", "認知症専門ケア研修｜委員会", "認知症ケア委員会。年度末の記録確認。"),
    ("3月", "2027-03-24", "認知症専門ケア研修｜研修", "認知症ケア研修の年度末振り返り。"),
]
st.set_page_config(page_title=APP_TITLE, layout="wide")


def safe_filename(name: str) -> str:
    name = str(name or "file").strip()
    name = re.sub(r'[\\/:*?"<>|]+', "_", name)
    name = re.sub(r"\s+", "_", name)
    return name[:120]


def connect_db():
    return sqlite3.connect(DB_PATH, check_same_thread=False)


def table_columns(conn, table_name):
    cur = conn.cursor()
    cur.execute(f"PRAGMA table_info({table_name})")
    return [row[1] for row in cur.fetchall()]


def add_column_if_missing(conn, table_name, column_name, column_type):
    if column_name not in table_columns(conn, table_name):
        conn.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_type}")


def init_db():
    conn = connect_db()
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS training_plan (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        theme TEXT UNIQUE,
        committee TEXT,
        frequency TEXT,
        required_count INTEGER DEFAULT 1,
        responsible_person TEXT,
        is_committee INTEGER DEFAULT 0,
        sort_order INTEGER DEFAULT 0
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS training_records (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        training_date TEXT,
        theme TEXT,
        staff TEXT,
        participants TEXT,
        record_link TEXT,
        memo TEXT,
        case_title TEXT,
        case_summary TEXT,
        staff_report_comment TEXT,
        created_at TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS training_schedule (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scheduled_date TEXT,
        scheduled_month TEXT,
        theme TEXT,
        staff TEXT,
        place TEXT,
        target_staff TEXT,
        memo TEXT,
        status TEXT DEFAULT '予定',
        created_at TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS training_attachments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        record_id INTEGER,
        attachment_type TEXT,
        original_filename TEXT,
        saved_path TEXT,
        file_note TEXT,
        uploaded_at TEXT
    )
    """)

    add_column_if_missing(conn, "training_plan", "responsible_person", "TEXT")
    add_column_if_missing(conn, "training_plan", "is_committee", "INTEGER DEFAULT 0")
    add_column_if_missing(conn, "training_plan", "sort_order", "INTEGER DEFAULT 0")

    for table, cols in {
        "training_schedule": [
            ("scheduled_date", "TEXT"),
            ("scheduled_month", "TEXT"),
            ("theme", "TEXT"),
            ("staff", "TEXT"),
            ("place", "TEXT"),
            ("target_staff", "TEXT"),
            ("memo", "TEXT"),
            ("status", "TEXT DEFAULT '予定'"),
            ("created_at", "TEXT"),
        ],
        "training_records": [
            ("training_date", "TEXT"),
            ("theme", "TEXT"),
            ("staff", "TEXT"),
            ("participants", "TEXT"),
            ("record_link", "TEXT"),
            ("memo", "TEXT"),
            ("case_title", "TEXT"),
            ("case_summary", "TEXT"),
            ("staff_report_comment", "TEXT"),
            ("created_at", "TEXT"),
        ],
        "training_attachments": [
            ("record_id", "INTEGER"),
            ("attachment_type", "TEXT"),
            ("original_filename", "TEXT"),
            ("saved_path", "TEXT"),
            ("file_note", "TEXT"),
            ("uploaded_at", "TEXT"),
        ],
    }.items():
        for col, typ in cols:
            add_column_if_missing(conn, table, col, typ)

    current_master_themes = [x[0] for x in TRAINING_MASTER]
    replaced_old_themes = [
        "感染症の予防", "身体拘束適正化", "虐待防止", "業務継続計画（感染症）",
        "業務継続計画（自然災害）", "認知症専門ケア研修"
    ]
    for old_theme in replaced_old_themes:
        if old_theme not in current_master_themes:
            cur.execute("DELETE FROM training_plan WHERE theme=?", (old_theme,))

    for sort_order, (theme, committee, frequency, required_count, is_committee) in enumerate(TRAINING_MASTER, start=1):
        cur.execute("""
        INSERT INTO training_plan(theme, committee, frequency, required_count, is_committee, sort_order)
        VALUES (?, ?, ?, ?, ?, ?)
        ON CONFLICT(theme) DO UPDATE SET
            committee=excluded.committee,
            frequency=excluded.frequency,
            required_count=excluded.required_count,
            is_committee=excluded.is_committee
        """, (theme, committee, frequency, required_count, is_committee, sort_order))

    # 既存DBで並び順が未設定の行には、現在のID順で並び順を付与する
    rows = cur.execute("SELECT id FROM training_plan WHERE sort_order IS NULL OR sort_order=0 ORDER BY id").fetchall()
    for idx, (row_id,) in enumerate(rows, start=1):
        cur.execute("UPDATE training_plan SET sort_order=? WHERE id=?", (idx, row_id))

    conn.commit()
    conn.close()


def read_sql(query, params=()):
    conn = connect_db()
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    return df


def execute_sql(query, params=(), return_lastrowid=False):
    conn = connect_db()
    cur = conn.cursor()
    cur.execute(query, params)
    conn.commit()
    last_id = cur.lastrowid
    conn.close()
    return last_id if return_lastrowid else None


def get_plan():
    return read_sql("SELECT * FROM training_plan ORDER BY sort_order ASC, id ASC")


def get_records():
    return read_sql("SELECT * FROM training_records ORDER BY training_date DESC, id DESC")


def get_schedule():
    return read_sql("SELECT * FROM training_schedule ORDER BY id ASC")


def get_attachments(record_id=None):
    if record_id is None:
        return read_sql("SELECT * FROM training_attachments ORDER BY uploaded_at DESC, id DESC")
    return read_sql("SELECT * FROM training_attachments WHERE record_id=? ORDER BY id DESC", (int(record_id),))


def fiscal_month_order(month_text):
    order = {m: i for i, m in enumerate(MONTHS, start=1)}
    return order.get(str(month_text), 99)


def get_row_month(row):
    month = str(row.get("scheduled_month", "") or "").strip()
    if month in MONTHS:
        return month
    dt = pd.to_datetime(row.get("scheduled_date", ""), errors="coerce")
    if pd.isna(dt):
        return ""
    return f"{dt.month}月"


def month_from_date_or_selected(date_text, selected_month=""):
    """予定日が入力されている場合は、日付から予定月を自動判定する。
    例：予定日が2026-07-10なら、予定月は7月として保存する。
    """
    dt = pd.to_datetime(date_text, errors="coerce")
    if not pd.isna(dt):
        detected_month = f"{dt.month}月"
        if detected_month in MONTHS:
            return detected_month
    selected_month = str(selected_month or "").strip()
    return selected_month if selected_month in MONTHS else ""


def progress_df():
    plan = get_plan()
    records = get_records()

    if records.empty or "theme" not in records.columns:
        counts = pd.DataFrame(columns=["theme", "done_count"])
    else:
        counts = records.groupby("theme").size().reset_index(name="done_count")

    df = plan.merge(counts, on="theme", how="left")
    df["done_count"] = df["done_count"].fillna(0).astype(int)
    df["remaining"] = (df["required_count"] - df["done_count"]).clip(lower=0)
    df["rate"] = df.apply(lambda r: min(r["done_count"] / r["required_count"], 1.0) if r["required_count"] else 0, axis=1)
    df["状況"] = df.apply(lambda r: "✅ 完了" if r["done_count"] >= r["required_count"] else ("⚠ 不足" if r["done_count"] > 0 else "❌ 未実施"), axis=1)
    return df


def monthly_schedule_matrix():
    plan = get_plan()
    schedule = get_schedule()

    plan_view = plan.copy()
    if "responsible_person" not in plan_view.columns:
        plan_view["responsible_person"] = ""
    plan_view["committee_display"] = plan_view.apply(
        lambda r: (str(r.get("committee", "") or "") + (f"\n担当：{str(r.get('responsible_person', '')).strip()}" if str(r.get("responsible_person", "") or "").strip() else "")).strip(),
        axis=1
    )
    matrix = plan_view[["theme", "committee_display", "frequency", "required_count"]].copy()
    matrix.columns = ["研修テーマ", "委員会（担当者名）", "頻度", "年間必要回数"]

    for m in MONTHS:
        matrix[m] = ""

    if schedule.empty:
        return matrix

    for col in ["scheduled_month", "scheduled_date", "theme", "staff", "status", "memo"]:
        if col not in schedule.columns:
            schedule[col] = ""

    for _, row in schedule.iterrows():
        month = get_row_month(row)
        theme = str(row.get("theme", "") or "").strip()

        if month not in MONTHS or theme == "":
            continue

        if theme not in matrix["研修テーマ"].values:
            continue

        dt = pd.to_datetime(row.get("scheduled_date", ""), errors="coerce")
        day_text = "" if pd.isna(dt) else f"{dt.day}日"
        status_text = str(row.get("status", "予定") or "予定").strip()
        staff_text = str(row.get("staff", "") or "").strip()
        staff_text = f"／{staff_text}" if staff_text else ""

        entry = f"{day_text} {status_text}{staff_text}".strip()

        idx = matrix.index[matrix["研修テーマ"] == theme][0]
        current = matrix.at[idx, month]
        matrix.at[idx, month] = entry if current == "" else current + "\n" + entry

    return matrix


def schedule_list_df():
    schedule = get_schedule()
    cols = ["id", "月", "予定日", "研修テーマ", "担当者", "場所", "対象者", "状態", "メモ"]

    if schedule.empty:
        return pd.DataFrame(columns=cols)

    for col in ["scheduled_month", "scheduled_date", "theme", "staff", "place", "target_staff", "status", "memo"]:
        if col not in schedule.columns:
            schedule[col] = ""

    df = schedule.copy()
    df["月"] = df.apply(get_row_month, axis=1)
    df["月順"] = df["月"].apply(fiscal_month_order)

    dt = pd.to_datetime(df["scheduled_date"], errors="coerce")
    df["予定日"] = dt.dt.strftime("%Y-%m-%d")
    df["予定日"] = df["予定日"].fillna("")

    df = df.rename(columns={
        "theme": "研修テーマ",
        "staff": "担当者",
        "place": "場所",
        "target_staff": "対象者",
        "status": "状態",
        "memo": "メモ"
    })

    for col in cols:
        if col not in df.columns:
            df[col] = ""

    return df.sort_values(["月順", "予定日", "研修テーマ"])[cols]


def attachments_summary_df():
    records = get_records()
    attachments = get_attachments()
    if records.empty:
        return pd.DataFrame(columns=["研修ID", "実施日", "研修テーマ", "事例資料数", "職員レポート数", "その他資料数", "コメント"])

    df = records[["id", "training_date", "theme", "staff_report_comment"]].copy()
    df.columns = ["研修ID", "実施日", "研修テーマ", "コメント"]

    for typ, label in [("事例資料", "事例資料数"), ("職員レポート", "職員レポート数"), ("その他", "その他資料数")]:
        if attachments.empty:
            counts = pd.DataFrame(columns=["record_id", label])
        else:
            counts = attachments[attachments["attachment_type"] == typ].groupby("record_id").size().reset_index(name=label)
        df = df.merge(counts, left_on="研修ID", right_on="record_id", how="left").drop(columns=["record_id"], errors="ignore")
        df[label] = df[label].fillna(0).astype(int)

    return df.sort_values(["実施日", "研修ID"], ascending=[False, False])


def save_uploaded_files(record_id, files, attachment_type, note=""):
    if not files:
        return 0

    saved_count = 0
    base_dir = CASE_DIR if attachment_type == "事例資料" else REPORT_DIR if attachment_type == "職員レポート" else UPLOAD_DIR / "others"
    base_dir.mkdir(parents=True, exist_ok=True)

    for file in files:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = safe_filename(file.name)
        saved_name = f"record{record_id}_{timestamp}_{filename}"
        saved_path = base_dir / saved_name

        with open(saved_path, "wb") as f:
            f.write(file.getbuffer())

        execute_sql("""
        INSERT INTO training_attachments(record_id, attachment_type, original_filename, saved_path, file_note, uploaded_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """, (int(record_id), attachment_type, file.name, str(saved_path), note, datetime.now().isoformat(timespec="seconds")))

        saved_count += 1

    return saved_count


def apply_recommended_schedule():
    """年間必要回数に合わせて、おすすめ予定を不足分だけ登録する。
    既に同じテーマの予定が必要回数以上ある場合は追加しない。
    """
    existing = get_schedule()
    plan = get_plan()
    required_map = {}
    if not plan.empty:
        required_map = dict(zip(plan["theme"], plan["required_count"]))

    added = 0
    existing_keys = set()
    existing_counts = {}

    if not existing.empty:
        existing_keys = set(zip(existing["scheduled_date"].fillna(""), existing["theme"].fillna("")))
        existing_counts = existing.groupby("theme").size().to_dict()

    for month, scheduled_date, theme, memo in RECOMMENDED_SCHEDULE_6_START:
        required_count = int(required_map.get(theme, 999))
        current_count = int(existing_counts.get(theme, 0))

        # すでに必要回数分の予定がある場合は、それ以上追加しない
        if current_count >= required_count:
            continue

        key = (scheduled_date, theme)
        if key in existing_keys:
            continue

        final_month = month_from_date_or_selected(scheduled_date, month)

        execute_sql("""
        INSERT INTO training_schedule(scheduled_date, scheduled_month, theme, staff, place, target_staff, memo, status, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            scheduled_date, final_month, theme, "",
            "フロア会議・委員会と同時実施",
            "全職員",
            memo,
            "予定",
            datetime.now().isoformat(timespec="seconds")
        ))

        existing_keys.add(key)
        existing_counts[theme] = current_count + 1
        added += 1

    return added


def audit_checklist_df():
    progress = progress_df()
    attach = attachments_summary_df()
    rows = []

    for _, r in progress.iterrows():
        theme = r["theme"]
        recs = get_records()
        theme_recs = recs[recs["theme"] == theme] if not recs.empty else pd.DataFrame()
        has_record = not theme_recs.empty
        has_case = False
        has_report = False
        has_comment = False

        if has_record and not attach.empty:
            ids = theme_recs["id"].tolist()
            a = attach[attach["研修ID"].isin(ids)]
            has_case = int(a["事例資料数"].sum()) > 0
            has_report = int(a["職員レポート数"].sum()) > 0
            has_comment = any(str(x).strip() for x in a["コメント"].fillna("").tolist())

        rows.append({
            "研修テーマ": theme,
            "必要回数": int(r["required_count"]),
            "実施数": int(r["done_count"]),
            "不足": int(r["remaining"]),
            "実施記録": "あり" if has_record else "なし",
            "事例資料": "あり" if has_case else "なし",
            "職員レポート": "あり" if has_report else "なし",
            "コメント保存": "あり" if has_comment else "なし",
            "監査準備状況": "✅ 準備OK" if (r["remaining"] == 0 and has_record and has_report) else "⚠ 確認必要"
        })

    return pd.DataFrame(rows)



def format_japanese_date(date_text, month_text=""):
    """予定日を短く表示する。日付がなければ月のみ。"""
    dt = pd.to_datetime(date_text, errors="coerce")
    if pd.isna(dt):
        return str(month_text or "")
    return f"{dt.month}/{dt.day}"



def normalize_plan_sort_order():
    """研修テーマの並び順を1,2,3...に整える。"""
    conn = connect_db()
    cur = conn.cursor()
    rows = cur.execute("SELECT id FROM training_plan ORDER BY sort_order ASC, id ASC").fetchall()
    for idx, (row_id,) in enumerate(rows, start=1):
        cur.execute("UPDATE training_plan SET sort_order=? WHERE id=?", (idx, row_id))
    conn.commit()
    conn.close()


def move_training_theme(target_id, direction):
    """研修テーマを上下に移動する。direction=-1で上、1で下。"""
    normalize_plan_sort_order()
    plan = get_plan()
    if plan.empty or target_id not in plan["id"].tolist():
        return False

    ids = plan["id"].tolist()
    current_index = ids.index(target_id)
    new_index = current_index + direction
    if new_index < 0 or new_index >= len(ids):
        return False

    ids[current_index], ids[new_index] = ids[new_index], ids[current_index]

    conn = connect_db()
    cur = conn.cursor()
    for idx, row_id in enumerate(ids, start=1):
        cur.execute("UPDATE training_plan SET sort_order=? WHERE id=?", (idx, int(row_id)))
    conn.commit()
    conn.close()
    return True


def set_training_theme_order(target_id, new_order):
    """指定した研修テーマを任意の表示順へ移動する。"""
    normalize_plan_sort_order()
    plan = get_plan()
    if plan.empty or target_id not in plan["id"].tolist():
        return False

    ids = plan["id"].tolist()
    ids.remove(target_id)
    insert_index = max(0, min(int(new_order) - 1, len(ids)))
    ids.insert(insert_index, target_id)

    conn = connect_db()
    cur = conn.cursor()
    for idx, row_id in enumerate(ids, start=1):
        cur.execute("UPDATE training_plan SET sort_order=? WHERE id=?", (idx, int(row_id)))
    conn.commit()
    conn.close()
    return True


def create_printable_annual_calendar_excel():
    """A4横1枚印刷を想定した年間研修スケジュールExcelを作成する。"""
    output = AUDIT_DIR / f"A4年間研修スケジュール_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    plan = get_plan()
    schedule = get_schedule()

    wb = Workbook()
    ws = wb.active
    ws.title = "A4年間スケジュール"

    # Page setup: A4 landscape, fit to one page
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1

    title_fill = PatternFill("solid", fgColor="D9EAD3")
    header_fill = PatternFill("solid", fgColor="EAF4E2")
    month_fill = PatternFill("solid", fgColor="DDEBF7")
    done_fill = PatternFill("solid", fgColor="D9EAD3")
    warn_fill = PatternFill("solid", fgColor="FCE4D6")
    plan_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:P1")
    ws["A1"] = "年間研修実施予定カレンダー（A4印刷用）"
    ws["A1"].font = Font(size=18, bold=True, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = title_fill
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:P2")
    ws["A2"] = "4月〜翌3月の研修予定を1枚で確認できます。予定変更・実施状況はシステム上で更新してください。"
    ws["A2"].font = Font(size=9, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    headers = ["研修テーマ", "委員会\n担当者", "必要"] + MONTHS + ["進捗"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = Font(size=9, bold=True, color="1F1F1F")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = month_fill if h in MONTHS else header_fill
        cell.border = border

    # Build schedule mapping
    sched_map = {}
    if not schedule.empty:
        for _, row in schedule.iterrows():
            theme = str(row.get("theme", "") or "").strip()
            month = get_row_month(row)
            if not theme or month not in MONTHS:
                continue
            status = str(row.get("status", "予定") or "予定")
            date_short = format_japanese_date(row.get("scheduled_date", ""), month)
            staff = str(row.get("staff", "") or "").strip()
            text = f"{date_short}\n{status}"
            if staff:
                text += f"\n{staff}"
            sched_map.setdefault((theme, month), []).append(text)

    progress = progress_df()
    progress_map = {}
    if not progress.empty:
        for _, r in progress.iterrows():
            progress_map[str(r["theme"])] = f'{int(r["done_count"])}/{int(r["required_count"])}'

    start_row = 4
    for r_idx, (_, p) in enumerate(plan.iterrows(), start=start_row):
        theme = str(p["theme"])
        committee_text = str(p.get("committee", "") or "")
        responsible_text = str(p.get("responsible_person", "") or "").strip()
        if responsible_text:
            committee_text = (committee_text + "\n担当：" + responsible_text).strip()
        values = [theme, committee_text, int(p.get("required_count", 0))]
        for m in MONTHS:
            values.append("\n---\n".join(sched_map.get((theme, m), [])))
        values.append(progress_map.get(theme, f'0/{int(p.get("required_count", 0))}'))

        for c_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if c_idx >= 3 else "left", vertical="center", wrap_text=True)
            cell.font = Font(size=8 if c_idx >= 4 else 9)
            if c_idx >= 4 and c_idx <= 15 and str(v).strip():
                cell.fill = done_fill if "実施済" in str(v) else plan_fill
            if c_idx == 16:
                if str(v).split("/")[0] == str(v).split("/")[-1]:
                    cell.fill = done_fill
                else:
                    cell.fill = warn_fill

        ws.row_dimensions[r_idx].height = 42

    last_row = start_row + len(plan) - 1

    # Bottom legend
    legend_row = last_row + 2
    ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=16)
    ws.cell(legend_row, 1).value = "色の見方：黄色＝予定　緑＝実施済　赤系＝不足確認　　※A4横1枚印刷用に、文字は小さめに設定しています。"
    ws.cell(legend_row, 1).font = Font(size=9, color="666666")
    ws.cell(legend_row, 1).alignment = Alignment(horizontal="center")
    ws.cell(legend_row, 1).fill = PatternFill("solid", fgColor="F8F8F8")

    widths = {
        1: 18, 2: 14, 3: 6, 16: 7
    }
    for col in range(1, 17):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 8.5)

    ws.freeze_panes = "D4"
    ws.print_area = f"A1:P{legend_row}"
    ws.sheet_view.showGridLines = False

    wb.save(output)
    return output


def schedule_edit_ui(schedule, themes, key_prefix="schedule_edit"):
    """予定の変更・更新・削除を共通表示する。
    Streamlitは同じkeyの入力値を保持するため、選択IDごとにフォーム部品のkeyを変える。
    これにより、予定を選び直したときに月・日付・テーマ・担当者などが正しく切り替わる。
    """
    if schedule.empty:
        st.info("更新・削除できる予定はまだありません。")
        return

    list_df = schedule_list_df()
    st.dataframe(list_df, use_container_width=True, hide_index=True)

    options = []
    for _, r in list_df.iterrows():
        label = f'ID{r["id"]}｜{r["月"]}｜{r["予定日"]}｜{r["研修テーマ"]}｜{r["状態"]}'
        options.append((label, int(r["id"])))

    selected_label = st.selectbox(
        "変更・更新する予定を選択",
        [x[0] for x in options],
        key=f"{key_prefix}_select"
    )
    target_id = dict(options)[selected_label]
    selected = schedule[schedule["id"] == target_id]

    if selected.empty:
        st.info("予定が見つかりません。")
        return

    row = selected.iloc[0]
    form_key = f"{key_prefix}_{target_id}"
    with st.form(f"{form_key}_form"):
        default_month = get_row_month(row)
        if default_month not in MONTHS:
            default_month = "4月"

        new_month = st.selectbox("予定月", MONTHS, index=MONTHS.index(default_month), key=f"{form_key}_month")
        has_date = bool(str(row.get("scheduled_date", "") or "").strip())
        use_date_edit = st.checkbox("具体的な予定日も入力する", value=has_date, key=f"{form_key}_use_date")

        new_date = ""
        if use_date_edit:
            default_date = pd.to_datetime(row.get("scheduled_date", ""), errors="coerce")
            target_year = 2026 if MONTH_NUM[new_month] >= 4 else 2027
            target_month = MONTH_NUM[new_month]
            if pd.isna(default_date) or default_date.month != target_month or default_date.year != target_year:
                default_date = pd.Timestamp(date(target_year, target_month, 1))
            # 予定月を変更したとき、日付欄もその月の1日に切り替わるようにする
            new_date = st.date_input("予定日", value=default_date.date(), key=f"{form_key}_date_{new_month}").isoformat()

        current_theme = str(row.get("theme", "") or "")
        new_theme = st.selectbox("研修テーマ", themes, index=themes.index(current_theme) if current_theme in themes else 0, key=f"{form_key}_theme")
        new_staff = st.text_input("担当者", value=row.get("staff", "") or "", key=f"{form_key}_staff")
        new_place = st.text_input("場所・実施方法", value=row.get("place", "") or "", key=f"{form_key}_place")
        new_target = st.text_input("対象者", value=row.get("target_staff", "") or "", key=f"{form_key}_target")

        status_options = ["予定", "延期", "中止", "実施待ち", "実施済"]
        current_status = str(row.get("status", "") or "予定")
        new_status = st.selectbox("状態", status_options, index=status_options.index(current_status) if current_status in status_options else 0, key=f"{form_key}_status")
        new_memo = st.text_area("メモ", value=row.get("memo", "") or "", key=f"{form_key}_memo")

        col1, col2 = st.columns(2)
        update_btn = col1.form_submit_button("更新する")
        delete_btn = col2.form_submit_button("削除する")

    if update_btn:
        final_month = month_from_date_or_selected(new_date, new_month)
        execute_sql("""
        UPDATE training_schedule
        SET scheduled_date=?, scheduled_month=?, theme=?, staff=?, place=?, target_staff=?, status=?, memo=?
        WHERE id=?
        """, (new_date, final_month, new_theme, new_staff, new_place, new_target, new_status, new_memo, int(target_id)))
        st.success(f"予定を更新しました。表示月も「{final_month}」に反映しました。")
        st.rerun()

    if delete_btn:
        execute_sql("DELETE FROM training_schedule WHERE id=?", (int(target_id),))
        st.warning("予定を削除しました。")
        st.rerun()


def create_audit_excel():
    output = AUDIT_DIR / f"監査ファイル_年間研修管理_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    records = get_records()
    attachments = get_attachments()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        monthly_schedule_matrix().to_excel(writer, sheet_name="01_年間予定表", index=False)
        schedule_list_df().to_excel(writer, sheet_name="02_予定一覧", index=False)
        progress_df().to_excel(writer, sheet_name="03_進捗一覧", index=False)
        records.to_excel(writer, sheet_name="04_研修実施記録", index=False)
        audit_checklist_df().to_excel(writer, sheet_name="05_監査チェックリスト", index=False)
        monthly_schedule_matrix().to_excel(writer, sheet_name="06_年間予定_一覧形式", index=False)
        attachments_summary_df().to_excel(writer, sheet_name="07_添付状況一覧", index=False)
        attachments.to_excel(writer, sheet_name="08_添付ファイル台帳", index=False)

    return output


def create_audit_zip():
    excel_path = create_audit_excel()
    printable_path = create_printable_annual_calendar_excel()
    zip_path = AUDIT_DIR / f"監査ファイル一式_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        z.write(excel_path, arcname=excel_path.name)
        z.write(printable_path, arcname=printable_path.name)

        for folder in [CASE_DIR, REPORT_DIR, UPLOAD_DIR / "others"]:
            if folder.exists():
                for file in folder.rglob("*"):
                    if file.is_file():
                        z.write(file, arcname=str(file))

    return zip_path


init_db()

st.title(APP_TITLE)
st.caption("年間研修の予定・実施記録・事例資料・職員レポート添付・監査ファイル出力を一体管理します。")

menu = st.sidebar.radio(
    "メニュー",
    [
        "管理者ダッシュボード",
        "年間実施予定カレンダー",
        "A4年間スケジュール印刷",
        "おすすめスケジュール登録",
        "研修予定登録",
        "研修実施入力・資料添付",
        "研修記録一覧・更新削除",
        "添付資料・レポート管理",
        "監査ファイル自動生成",
        "研修計画管理",
        "Excel出力"
    ]
)

if menu == "管理者ダッシュボード":
    st.header("管理者ダッシュボード")

    df = progress_df()
    total_required = int(df["required_count"].sum()) if not df.empty else 0
    total_done = int(df["done_count"].sum()) if not df.empty else 0
    total_rate = total_done / total_required if total_required else 0
    attach_df = attachments_summary_df()

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("年間必要回数", total_required)
    c2.metric("実施済み回数", total_done)
    c3.metric("年間実施率", f"{total_rate:.0%}")
    c4.metric("未実施・不足項目", int((df["done_count"] < df["required_count"]).sum()))

    c5, c6, c7 = st.columns(3)
    c5.metric("事例資料あり研修", int((attach_df["事例資料数"] > 0).sum()) if not attach_df.empty and "事例資料数" in attach_df.columns else 0)
    c6.metric("職員レポートあり研修", int((attach_df["職員レポート数"] > 0).sum()) if not attach_df.empty and "職員レポート数" in attach_df.columns else 0)
    c7.metric("コメント保存あり研修", int(attach_df["コメント"].fillna("").astype(str).str.strip().ne("").sum()) if not attach_df.empty and "コメント" in attach_df.columns else 0)

    st.subheader("年間実施予定（月単位）")
    st.dataframe(monthly_schedule_matrix(), use_container_width=True, hide_index=True, height=360)

    st.subheader("監査チェックリスト")
    st.dataframe(audit_checklist_df(), use_container_width=True, hide_index=True)

    st.subheader("研修進捗一覧")
    if "responsible_person" not in df.columns:
        df["responsible_person"] = ""
    view = df[["theme", "committee", "responsible_person", "frequency", "required_count", "done_count", "remaining", "rate", "状況"]].copy()
    view.columns = ["研修テーマ", "委員会", "担当者名", "頻度", "必要回数", "実施数", "残り", "実施率", "状況"]
    st.dataframe(
        view,
        use_container_width=True,
        hide_index=True,
        column_config={"実施率": st.column_config.ProgressColumn("実施率", min_value=0, max_value=1, format="%.0f%%")}
    )

    st.subheader("年間実施率グラフ")
    chart_df = df[["theme", "rate"]].copy()
    chart_df["rate_percent"] = chart_df["rate"] * 100
    chart = (
        alt.Chart(chart_df)
        .mark_bar()
        .encode(
            x=alt.X("rate_percent:Q", title="実施率（%）", scale=alt.Scale(domain=[0, 100])),
            y=alt.Y("theme:N", title="研修テーマ", sort="-x"),
            tooltip=["theme", alt.Tooltip("rate_percent:Q", format=".0f")]
        )
        .properties(height=380)
    )
    st.altair_chart(chart, use_container_width=True)

elif menu == "年間実施予定カレンダー":
    st.header("年間実施予定カレンダー")
    st.caption("研修テーマごとに、4月〜翌3月の予定を横並びで確認できます。予定の変更・更新もこの画面で行えます。")

    matrix = monthly_schedule_matrix()
    st.dataframe(matrix, use_container_width=True, hide_index=True, height=500)

    st.subheader("月別の予定一覧")
    # 下の「変更・更新する予定」を選び直したとき、月別一覧の表示月もその予定月へ合わせる
    current_edit_label = st.session_state.get("calendar_edit_select", "")
    if current_edit_label and st.session_state.get("_last_calendar_edit_select") != current_edit_label:
        parts = str(current_edit_label).split("｜")
        if len(parts) >= 2 and parts[1] in MONTHS:
            st.session_state["calendar_display_month"] = parts[1]
        st.session_state["_last_calendar_edit_select"] = current_edit_label

    selected_month = st.selectbox("表示する月", MONTHS, key="calendar_display_month")
    list_df = schedule_list_df()
    month_df = list_df[list_df["月"] == selected_month].drop(columns=["id"], errors="ignore")

    if month_df.empty:
        st.info(f"{selected_month}の予定はまだ登録されていません。")
    else:
        st.dataframe(month_df, use_container_width=True, hide_index=True)

    st.divider()
    st.subheader("予定の変更・更新・削除")
    st.caption("下の一覧から予定を選択して、月・日付・研修テーマ・担当者・状態などを変更できます。")
    plan = get_plan()
    themes = plan["theme"].tolist()
    schedule_edit_ui(get_schedule(), themes, key_prefix="calendar_edit")

elif menu == "A4年間スケジュール印刷":
    st.header("A4年間スケジュール印刷")
    st.caption("A4横1枚で印刷しやすい年間研修スケジュールExcelを作成します。")

    st.subheader("印刷イメージ用データ")
    st.dataframe(monthly_schedule_matrix(), use_container_width=True, hide_index=True, height=420)

    if st.button("A4年間スケジュールExcelを作成する"):
        output = create_printable_annual_calendar_excel()
        with open(output, "rb") as f:
            st.download_button(
                label="A4年間スケジュールExcelをダウンロード",
                data=f,
                file_name=output.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

elif menu == "おすすめスケジュール登録":
    st.header("6月開始おすすめスケジュール登録")
    st.caption("監査前に慌てないよう、6月開始で分散配置したおすすめ予定を一括登録します。")

    preview = pd.DataFrame(RECOMMENDED_SCHEDULE_6_START, columns=["月", "予定日", "研修テーマ", "狙い・メモ"])
    st.dataframe(preview, use_container_width=True, hide_index=True)

    if st.button("このおすすめスケジュールを一括登録する"):
        added = apply_recommended_schedule()
        st.success(f"{added}件の予定を登録しました。既に同じ日付・テーマの予定があるものは重複登録していません。")

elif menu == "研修予定登録":
    st.header("研修予定登録")
    st.caption("日付未定でも、予定月だけで登録できます。")

    plan = get_plan()
    themes = plan["theme"].tolist()

    with st.form("schedule_form"):
        scheduled_month = st.selectbox("予定月", MONTHS)
        use_date = st.checkbox("具体的な予定日も入力する", value=False)

        scheduled_date = ""
        if use_date:
            year = 2026 if MONTH_NUM[scheduled_month] >= 4 else 2027
            scheduled_date = st.date_input("予定日", value=date(year, MONTH_NUM[scheduled_month], 1)).isoformat()

        theme = st.selectbox("研修テーマ", themes)
        staff = st.text_input("担当者")
        place = st.text_input("場所・実施方法", placeholder="例：事務室／オンライン／フロア会議")
        target_staff = st.text_input("対象者", placeholder="例：全職員／夜勤者／新規採用者")
        status = st.selectbox("状態", ["予定", "延期", "中止", "実施待ち", "実施済"])
        memo = st.text_area("メモ")
        submitted = st.form_submit_button("予定を登録する")

    if submitted:
        final_month = month_from_date_or_selected(scheduled_date, scheduled_month)
        execute_sql("""
        INSERT INTO training_schedule(scheduled_date, scheduled_month, theme, staff, place, target_staff, memo, status, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (scheduled_date, final_month, theme, staff, place, target_staff, memo, status, datetime.now().isoformat(timespec="seconds")))
        st.success(f"研修予定を登録しました。予定月は「{final_month}」で保存しました。")

    st.subheader("予定の更新・削除")
    schedule = get_schedule()
    if schedule.empty:
        st.info("更新・削除できる予定はまだありません。")
    else:
        list_df = schedule_list_df()
        st.dataframe(list_df, use_container_width=True, hide_index=True)

        target_id = st.number_input("対象ID", min_value=1, step=1, key="schedule_id")
        selected = schedule[schedule["id"] == target_id]

        if selected.empty:
            st.info("上の一覧から更新・削除したいIDを入力してください。")
        else:
            row = selected.iloc[0]
            with st.form("edit_schedule"):
                default_month = get_row_month(row)
                if default_month not in MONTHS:
                    default_month = "4月"

                new_month = st.selectbox("予定月", MONTHS, index=MONTHS.index(default_month))
                has_date = bool(str(row.get("scheduled_date", "") or "").strip())
                use_date_edit = st.checkbox("具体的な予定日も入力する", value=has_date)

                new_date = ""
                if use_date_edit:
                    default_date = pd.to_datetime(row.get("scheduled_date", ""), errors="coerce")
                    if pd.isna(default_date):
                        year = 2026 if MONTH_NUM[new_month] >= 4 else 2027
                        default_date = pd.Timestamp(date(year, MONTH_NUM[new_month], 1))
                    new_date = st.date_input("予定日", value=default_date.date(), key="edit_schedule_date").isoformat()

                current_theme = str(row.get("theme", "") or "")
                new_theme = st.selectbox("研修テーマ", themes, index=themes.index(current_theme) if current_theme in themes else 0)
                new_staff = st.text_input("担当者", value=row.get("staff", "") or "")
                new_place = st.text_input("場所・実施方法", value=row.get("place", "") or "")
                new_target = st.text_input("対象者", value=row.get("target_staff", "") or "")

                status_options = ["予定", "延期", "中止", "実施待ち", "実施済"]
                current_status = str(row.get("status", "") or "予定")
                new_status = st.selectbox("状態", status_options, index=status_options.index(current_status) if current_status in status_options else 0)
                new_memo = st.text_area("メモ", value=row.get("memo", "") or "")

                col1, col2 = st.columns(2)
                update_btn = col1.form_submit_button("更新する")
                delete_btn = col2.form_submit_button("削除する")

            if update_btn:
                final_month = month_from_date_or_selected(new_date, new_month)
                execute_sql("""
                UPDATE training_schedule
                SET scheduled_date=?, scheduled_month=?, theme=?, staff=?, place=?, target_staff=?, status=?, memo=?
                WHERE id=?
                """, (new_date, final_month, new_theme, new_staff, new_place, new_target, new_status, new_memo, int(target_id)))
                st.success(f"予定を更新しました。表示月も「{final_month}」に反映しました。")
                st.rerun()

            if delete_btn:
                execute_sql("DELETE FROM training_schedule WHERE id=?", (int(target_id),))
                st.warning("予定を削除しました。")

elif menu == "研修実施入力・資料添付":
    st.header("研修実施入力・資料添付")
    st.caption("使用した事例資料、職員レポートExcel、コメントを研修記録に紐づけて保存できます。")

    plan = get_plan()
    themes = plan["theme"].tolist()

    with st.form("record_form"):
        training_date = st.date_input("実施日", value=date.today())
        theme = st.selectbox("研修名", themes)
        staff = st.text_input("担当者")
        participants = st.text_area("参加者")
        record_link = st.text_input("研修記録リンク（Google Drive、PDF、写真URLなど）")
        memo = st.text_area("備考・メモ")

        st.markdown("### 使用した事例")
        case_title = st.text_input("事例タイトル", placeholder="例：夜間トイレ移動時の転倒リスク事例")
        case_summary = st.text_area("事例内容メモ", placeholder="研修で使用した事例の要点を記録します。")
        case_files = st.file_uploader(
            "使用した事例ファイルを添付（Excel・Word・PDF・画像など）",
            type=["xlsx", "xls", "docx", "doc", "pdf", "png", "jpg", "jpeg", "txt"],
            accept_multiple_files=True
        )

        st.markdown("### 職員レポート")
        report_files = st.file_uploader(
            "職員レポートを添付（Excel推奨。Word・PDFも可）",
            type=["xlsx", "xls", "docx", "doc", "pdf", "txt"],
            accept_multiple_files=True
        )
        staff_report_comment = st.text_area("職員レポートへのコメント・確認メモ", placeholder="例：全職員提出済。転倒リスクへの理解が確認できた。")

        other_files = st.file_uploader(
            "その他添付（写真・議事録・配布資料など）",
            type=["xlsx", "xls", "docx", "doc", "pdf", "png", "jpg", "jpeg", "txt"],
            accept_multiple_files=True
        )

        submitted = st.form_submit_button("研修記録と添付資料を登録する")

    if submitted:
        record_id = execute_sql("""
        INSERT INTO training_records(training_date, theme, staff, participants, record_link, memo, case_title, case_summary, staff_report_comment, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            training_date.isoformat(), theme, staff, participants, record_link, memo,
            case_title, case_summary, staff_report_comment,
            datetime.now().isoformat(timespec="seconds")
        ), return_lastrowid=True)

        case_count = save_uploaded_files(record_id, case_files, "事例資料", case_title)
        report_count = save_uploaded_files(record_id, report_files, "職員レポート", staff_report_comment)
        other_count = save_uploaded_files(record_id, other_files, "その他", memo)

        execute_sql("""
        UPDATE training_schedule
        SET status='実施済'
        WHERE theme=? AND (scheduled_date=? OR scheduled_month=?)
        """, (theme, training_date.isoformat(), f"{training_date.month}月"))

        st.success(f"研修記録を登録しました。事例資料 {case_count}件、職員レポート {report_count}件、その他 {other_count}件を保存しました。")

elif menu == "研修記録一覧・更新削除":
    st.header("研修記録一覧・更新削除")
    records = get_records()

    if records.empty:
        st.info("まだ研修記録がありません。")
    else:
        st.dataframe(records, use_container_width=True, hide_index=True)

        st.subheader("記録の更新・削除")
        record_id = st.number_input("対象ID", min_value=1, step=1)
        selected = records[records["id"] == record_id]

        if selected.empty:
            st.info("上の一覧から更新・削除したいIDを入力してください。")
        else:
            row = selected.iloc[0]
            plan = get_plan()
            themes = plan["theme"].tolist()

            with st.form("edit_record"):
                new_date = st.date_input("実施日", value=pd.to_datetime(row["training_date"]).date())
                new_theme = st.selectbox("研修名", themes, index=themes.index(row["theme"]) if row["theme"] in themes else 0)
                new_staff = st.text_input("担当者", value=row.get("staff", "") or "")
                new_participants = st.text_area("参加者", value=row.get("participants", "") or "")
                new_link = st.text_input("研修記録リンク", value=row.get("record_link", "") or "")
                new_memo = st.text_area("備考・メモ", value=row.get("memo", "") or "")
                new_case_title = st.text_input("事例タイトル", value=row.get("case_title", "") or "")
                new_case_summary = st.text_area("事例内容メモ", value=row.get("case_summary", "") or "")
                new_report_comment = st.text_area("職員レポートへのコメント・確認メモ", value=row.get("staff_report_comment", "") or "")

                col1, col2 = st.columns(2)
                update_btn = col1.form_submit_button("更新する")
                delete_btn = col2.form_submit_button("削除する")

            if update_btn:
                execute_sql("""
                UPDATE training_records
                SET training_date=?, theme=?, staff=?, participants=?, record_link=?, memo=?, case_title=?, case_summary=?, staff_report_comment=?
                WHERE id=?
                """, (new_date.isoformat(), new_theme, new_staff, new_participants, new_link, new_memo, new_case_title, new_case_summary, new_report_comment, int(record_id)))
                st.success("更新しました。")

            if delete_btn:
                execute_sql("DELETE FROM training_records WHERE id=?", (int(record_id),))
                execute_sql("DELETE FROM training_attachments WHERE record_id=?", (int(record_id),))
                st.warning("記録と添付台帳を削除しました。保存済ファイル自体は監査用バックアップとしてフォルダに残ります。")

            st.subheader("この記録に紐づく添付資料")
            att = get_attachments(int(record_id))
            if att.empty:
                st.info("添付資料はありません。")
            else:
                st.dataframe(att, use_container_width=True, hide_index=True)

elif menu == "添付資料・レポート管理":
    st.header("添付資料・レポート管理")
    st.caption("研修ごとの事例資料・職員レポート提出状況を確認します。")

    summary = attachments_summary_df()
    st.subheader("添付状況サマリー")
    st.dataframe(summary, use_container_width=True, hide_index=True)

    st.subheader("添付ファイル台帳")
    att = get_attachments()
    if att.empty:
        st.info("添付ファイルはまだありません。")
    else:
        st.dataframe(att, use_container_width=True, hide_index=True)

        target_id = st.number_input("削除する添付ID", min_value=1, step=1)
        if st.button("添付台帳から削除する"):
            execute_sql("DELETE FROM training_attachments WHERE id=?", (int(target_id),))
            st.warning("添付台帳から削除しました。保存済ファイル自体はフォルダに残ります。")

elif menu == "監査ファイル自動生成":
    st.header("監査ファイル自動生成")
    st.caption("年間予定、進捗、実施記録、添付資料台帳、チェックリストを監査提出用にまとめます。")

    st.subheader("監査チェックリスト")
    st.dataframe(audit_checklist_df(), use_container_width=True, hide_index=True)

    col1, col2, col3 = st.columns(3)

    with col1:
        if st.button("監査用Excelを作成する"):
            output = create_audit_excel()
            with open(output, "rb") as f:
                st.download_button(
                    label="監査用Excelをダウンロード",
                    data=f,
                    file_name=output.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    with col2:
        if st.button("監査ファイル一式ZIPを作成する"):
            zip_path = create_audit_zip()
            with open(zip_path, "rb") as f:
                st.download_button(
                    label="監査ファイル一式ZIPをダウンロード",
                    data=f,
                    file_name=zip_path.name,
                    mime="application/zip"
                )

    with col3:
        if st.button("A4年間スケジュールを作成する"):
            output = create_printable_annual_calendar_excel()
            with open(output, "rb") as f:
                st.download_button(
                    label="A4年間スケジュールをダウンロード",
                    data=f,
                    file_name=output.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    st.info("ZIPには監査用Excel、A4年間スケジュール、保存済みの事例資料・職員レポート等の添付ファイルが含まれます。")

elif menu == "研修計画管理":
    st.header("研修計画管理")
    st.caption("写真の年間研修表に合わせ、委員会・研修・訓練を分けて管理します。委員会形式の項目は担当者名を登録できます。研修テーマの表示順も変更できます。")
    plan = get_plan()

    view = plan.copy()
    if "is_committee" not in view.columns:
        view["is_committee"] = 0
    if "responsible_person" not in view.columns:
        view["responsible_person"] = ""
    if "sort_order" not in view.columns:
        view["sort_order"] = range(1, len(view) + 1)
    view["表示順"] = view["sort_order"].fillna(0).astype(int)
    view["委員会形式"] = view["is_committee"].apply(lambda x: "対象" if int(x or 0) == 1 else "")
    view = view.rename(columns={
        "theme": "研修テーマ",
        "committee": "委員会",
        "responsible_person": "担当者名",
        "frequency": "頻度",
        "required_count": "必要回数"
    })
    st.dataframe(view[["表示順", "研修テーマ", "委員会", "担当者名", "頻度", "必要回数", "委員会形式"]], use_container_width=True, hide_index=True)

    st.subheader("研修テーマの並び順変更")
    st.caption("ここで変更した順番は、管理者ダッシュボード、年間カレンダー、A4年間スケジュール、Excel出力に反映されます。")
    if plan.empty:
        st.info("並び順を変更できる研修テーマがありません。")
    else:
        order_options = []
        for _, r in plan.iterrows():
            label = f'{int(r.get("sort_order", 0) or 0)}｜{r["theme"]}'
            order_options.append((label, int(r["id"])))
        selected_order_label = st.selectbox("並び順を変更する研修テーマ", [x[0] for x in order_options], key="sort_select_theme")
        selected_order_id = dict(order_options)[selected_order_label]
        col_up, col_down, col_set = st.columns([1, 1, 2])
        if col_up.button("上へ移動", key="move_up"):
            if move_training_theme(selected_order_id, -1):
                st.success("1つ上へ移動しました。画面を再読み込みすると反映されます。")
            else:
                st.info("これ以上上には移動できません。")
        if col_down.button("下へ移動", key="move_down"):
            if move_training_theme(selected_order_id, 1):
                st.success("1つ下へ移動しました。画面を再読み込みすると反映されます。")
            else:
                st.info("これ以上下には移動できません。")
        with col_set:
            new_order = st.number_input("指定した順番へ移動", min_value=1, max_value=max(len(plan), 1), value=1, step=1, key="new_sort_order")
            if st.button("この順番に移動", key="set_sort_order"):
                if set_training_theme_order(selected_order_id, int(new_order)):
                    st.success(f"{int(new_order)}番目へ移動しました。画面を再読み込みすると反映されます。")

    st.divider()
    st.subheader("委員会担当者の登録・変更")
    committee_plan = plan[plan.get("is_committee", 0).fillna(0).astype(int) == 1].copy() if not plan.empty else pd.DataFrame()
    if committee_plan.empty:
        st.info("担当者を入力できる委員会形式の項目がありません。")
    else:
        options = committee_plan["theme"].tolist()
        selected_theme = st.selectbox("担当者を設定する委員会項目", options)
        selected_row = committee_plan[committee_plan["theme"] == selected_theme].iloc[0]
        with st.form("committee_responsible_form"):
            new_committee = st.text_input("委員会名", value=selected_row.get("committee", "") or "")
            new_responsible = st.text_input("担当者名", value=selected_row.get("responsible_person", "") or "", placeholder="例：渕本／管理者／感染対策担当")
            submitted = st.form_submit_button("担当者を保存する")
        if submitted:
            execute_sql("UPDATE training_plan SET committee=?, responsible_person=? WHERE theme=?", (new_committee, new_responsible, selected_theme))
            st.success("委員会名・担当者名を保存しました。年間予定表とA4印刷にも反映されます。")

elif menu == "Excel出力":
    st.header("Excel出力")

    output = Path("training_export.xlsx")
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        monthly_schedule_matrix().to_excel(writer, sheet_name="年間予定表_月単位", index=False)
        schedule_list_df().to_excel(writer, sheet_name="予定一覧", index=False)
        progress_df().to_excel(writer, sheet_name="進捗一覧", index=False)
        get_records().to_excel(writer, sheet_name="研修記録", index=False)
        attachments_summary_df().to_excel(writer, sheet_name="添付状況一覧", index=False)
        get_attachments().to_excel(writer, sheet_name="添付ファイル台帳", index=False)
        audit_checklist_df().to_excel(writer, sheet_name="監査チェックリスト", index=False)
        monthly_schedule_matrix().to_excel(writer, sheet_name="A4予定表用データ", index=False)

    with open(output, "rb") as f:
        st.download_button(
            label="Excelをダウンロード",
            data=f,
            file_name="年間研修管理_添付資料対応版.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

st.caption("Ver1.9.4：おすすめスケジュールの予定回数を年間必要回数に一致。既存予定がある場合は不足分のみ追加。")